# 여러 소스들을 임포트 ㅇㅅㅇ
import discord
import time
import asyncio
from datetime import datetime

from discord.utils import get
from discord.ext import commands
import openpyxl
import random

timervar = True
finish1 = False
timervar1 = True
client = discord.Client()


@client.event
async def on_message(message):
    if message.author == client.user:
        return


@client.event
async def on_ready():
    print(client.user.id)
    print("준비 완료!")
    game = discord.Game("pg!도움말")
    await client.change_presence(status=discord.Status.online, activity=game)
    print(datetime.today().month)
    print(datetime.today().day)


@client.event
async def on_message(message):
    if message.author == client.user:
        return
    if message.content.startswith("pg!도움말"):
        # 도움말 엠베드 설정
        embed = discord.Embed(title="먀우 명령어 도움말", description="걍 빨리 확인하고 입력하라능", color=0xff0000)
        embed.set_author(name="먀우 by 팀 먀우", url="https://playgroundkkutu.kro.kr",
                         icon_url="https://media.discordapp.net/attachments/674782331797504000/707760894637113345/556d695e84f57c9a.jpg")
        embed.set_thumbnail(url="https://media.discordapp.net/attachments/674782331797504000/708064796070772837/KakaoTalk_20200402_205852061.jpg")
        embed.add_field(name="pg!기본세팅", value="기본적으로 세팅해야 할 것을 세팅합니다. 관리자용.")
        embed.add_field(name="pg!역할지급 구독자", value="공지가 등록되었을 떄 맨션을 받을 수 있는 구독자 권한을 드립니다", inline=False)
        embed.add_field(name="pg!역할지급 19", value="19금 방에 들어갈 수 있는 [ 19 ] 권한을 드립니다. 쓰고 후회하지 마세요!", inline=False)
        #embed.add_field(name="pg!디엠건의", value="DM으로 놀고 있는 관리자를 깨웁니다", inline=False)
        #embed.add_field(name="pg!디엠쿠키", value="DM으로 놀고 있는 그쿠를 깨웁니다", inline=False)
        embed.add_field(name="pg!출석정보만들기", value="출석 정보를 만듭니다. 출석체크를 하기 전 해 주세요!(최초 1번만 실행하면 됨)", inline=False)
        embed.add_field(name="pg!출첵", value="자신의 이름으로 출석체크합니다.")
        embed.add_field(name="pg!출순 [닉네임]", value="[닉네임] 유저의 출석 정보를 봅니다.")
        embed.add_field(name="더 많은 명령어는?", value="60초안에 👍 이모지를 달아주시면 다음 리스트를 로딩합니다", inline=True)
        # cody by beta(beta에서 가져왔으면 걍 삭제 하지 왜납뒀냐 ㅇㅅㅇ)
        # embed.add_field(name="pg!역할제거 19", value="[ 19 ] 권한을 자신에게서 뺍니다", inline=False)
        # embed.add_field(name="pg!역할제거 구독자", value="구독자 권한을 자신에게서 뺍니다", inline=False)
        # embed.add_field(name="pg!도배", value="글자를 덕덕 알고리즘으로 극대화하여 도배합니다. 도배방에서 사용해주세요!", inline=False)
        embed.set_footer(text="먀우 많이 사랑해주세요!")
        # 출력 엠베드
        await message.channel.send(embed=embed)
        channel = message.channel

        def check(reaction, user):
            return user == message.author and str(reaction.emoji) == '👍'

        try:
            reaction, user = await client.wait_for('reaction_add', timeout=60.0, check=check)
        except asyncio.TimeoutError:
            await channel.send('👎타임 오버! 다시 해주세요..')
        else:
            embed1 = discord.Embed(title="먀우 명령어 도움말", description="걍 빨리 확인하고 입력하라능", color=0x0000ff)
            embed1.set_author(name="먀우 by 팀 먀우", url="https://playgroundkkutu.kro.kr",
                              icon_url="https://media.discordapp.net/attachments/674782331797504000/707760894637113345/556d695e84f57c9a.jpg")
            embed1.set_thumbnail(url="https://media.discordapp.net/attachments/674782331797504000/708064796070772837/KakaoTalk_20200402_205852061.jpg")
            embed1.add_field(name="pg!역할제거 19", value="[ 19 ] 권한을 자신에게서 뺍니다", inline=False)
            embed1.add_field(name="pg!역할제거 구독자", value="구독자 권한을 자신에게서 뺍니다", inline=False)
            embed1.add_field(name="pg!도배", value="글자를 덕덕 알고리즘으로 극대화하여 도배합니다. 도배방에서 사용해주세요!", inline=False)
            embed1.add_field(name="pg!파도", value="글자를 파도로 만듭니다. 도배방에서 사용해주세요!", inline=False)
            embed1.add_field(name="pg!가위바위종이 [가위 또는 바위 또는 종이]", value="먀우랑 하는 가위바위종이 게임입니다.", inline=False)
            embed1.add_field(name="pg!먀우수[숫자1]부터[숫자2]까지", value="숫자1(한자리수만)부터 숫자2(한자리수만)까지 먀우가 말합니다.", inline=False)
            embed1.add_field(name="pg!배워 [항목1] [항목2]", value="pg!배운거말해 [항목1] 을 입력했을 때 먀우가 [항목2]를 말합니다.", inline=False)
            embed1.add_field(name="pg!배운거말해", value="pg!배워 의 설명을 참고하세요!")
            embed1.add_field(name="pg!수정 [항목1] [항목2]", value="자신이 쓴 [항목 1]에 대한 설명을 [항목 2]로 바꿉니다.")
            embed1.add_field(name="pg!강템생성 [항목1]", value="[항목1]이라는 이름으로 강화 아이템을 생성합니다.", inline=False)
            embed1.add_field(name="pg!강화 [항목1]", value="[항목1]이라는 강화 아이템을 강화합니다. 행운을 빌어요!")
            embed1.add_field(name="pg!견적서 [항목1]", value="자신이 생성한 [항목1]이라는 강화 아이템의 견적서를 보여줍니다.")
            embed1.add_field(name="pg!운겜생성 [내 아바타 이름]", value="[내 아바타 이름]이라는 이름으로 내 아바타를 참가시킨 운빨겜을 생성합니다.")
            embed1.set_footer(text="먀우 많이 사랑해주세요!")
            await message.channel.send(embed=embed1)
    if message.content.startswith("pg!먀우수"):
        try:
            num1 = int(message.content[6])
            num2 = int(message.content[9])
        except ValueError:
            await message.channel.send("어.. 이상해요.. 제대로 입력하신 거 맞으세요?")
        # 걍 else로 처리하면 될것을
        if num1<num2:
            while num1<=num2:
                await message.channel.send(num1)
                num1 = num1 + 1
            msg = await message.channel.send("어때? 수학실력 쩔지?")
            time.sleep(1)
            await msg.edit(content="판사님 저는 아무 말도 하지 않았습니다")
        # 하 코드 효율 개쓰레기로 만들었네 ㅇㅅㅌ
        if num1>num2:
            while num1>=num2:
                await message.channel.send(num1)
                num1 = num1 - 1
            msg = await message.channel.send("어때? 수학실력 쩔지?")
            time.sleep(1)
            await msg.edit(content="판사님 저는 아무 말도 하지 않았습니다")
        # 덕봇수 수정 끝
    if message.content.startswith("pg!기본세팅"):
        if message.channel.permissions_for(message.author).administrator:
            learn = message.content.split(" ")
            try:
                if learn[1] == "승인":
                    await message.channel.send("안녕하세요! 팀 먀우에서 제작한 먀우봇을 서버에 넣어주셔서 감사합니다.")
                    time.sleep(2)
                    await message.channel.send("관리자 분은 **지금부터 제가 말하는 것을 따라해주세요.**")
                    time.sleep(1)
                    await message.channel.send("| 🎁구독자 | 라는 이름으로 역할을 만들어주시고, 앞으로는 공지에 everyone 대신 | 🎁구독자 |를 맨션해주세요.")
                    time.sleep(1)
                    await message.channel.send("| 19 [ 19 ] | 라는 이름으로 역할을 만들어주시고, 19금방을 | 19 [ 19 ] | 역할만 들어갈 수 있게 설정해주세요.")
                    time.sleep(2)
                    await message.channel.send("설정이 완료되었습니다! 먀우봇을 서버에 넣어주셔서 감사합니다.")
            except IndexError:
                await message.channel.send("`pg!기본세팅 승인`을 입력하시면 기본 채팅방에서 일부 명령어 사용 가능 허락, 간접적인 노리터끄투 홍보에 동의한 것으로 간주합니다.")
    if message.content.startswith("pg!디엠쿠키"):
        id = message.author.id
        me = get(message.guild.members, id=653075791814590487)
        channel = await me.create_dm()
        await channel.send("건의가 도착했습니다.")
        await message.channel.send("<@" + str(id) + "> 님,정상적으로 전송되었습니다!")
    if message.content.startswith("pg!경고"):
        if message.channel.permissions_for(message.author).administrator:
            learn = message.content.split(" ")
            id = learn[1]
            me = get(message.guild.members, id=int(id))
            channel = await me.create_dm()
            await channel.send("경고가 도착했습니다. 자세한 것은 관리자에게 물어보세요.")
    if message.content.startswith("pg!디엠건의"):
        id = message.author.id
        me = get(message.guild.members, id=488670402118156298)
        channel = await me.create_dm()
        await channel.send("건의가 도착했습니다.")
        await message.channel.send("<@" + str(id) + "> 님,정상적으로 전송되었습니다!")
    if message.content.startswith("pg!역할지급 구독자"):
        id = message.author.id
        role = discord.utils.get(message.guild.roles, name="| 🎁구독자 |")
        await message.author.add_roles(role)
        await message.channel.send("<@" + str(id) + "> 님,성공적으로 지급되었습니다!")

#출첵 관련 코드
    if message.content.startswith("pg!출첵"):
        file = openpyxl.load_workbook('\sheets\chullcheck.xlsx')
        sheet = file.active
        id = message.author.id
        gid = message.author.guild.name
        for i in range(1, 151):
            if sheet["A" + str(i)].value == message.author.name and sheet["E" + str(i)].value == str(gid):
                dayday = str(datetime.today().month) + str(datetime.today().day)
                print(dayday)
                if sheet["C" + str(i)].value == int(dayday):
                    await message.channel.send("<@" + str(id) + "> 님, 출석 체크는 하루에 한 번만 가능합니다.")
                    break
                else:
                    sheet["C" + str(i)].value = int(dayday)
                if sheet["B" + str(i)].value == "-":
                    sheet["B" + str(i)].value = 1
                else:
                    sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 1
                if sheet["B" + str(i)].value == 100:
                    me = get(message.guild.members, id=int(message.author.id))
                    channel = await me.create_dm()
                    await channel.send(message.author.name + " 님 출첵 100회 달성. 3000핑 지급 바람.")
                    await message.channel.send("<@" + str(id) + "> 님, 출첵 100회를 달성했습니다! 곧 유저분의 끄투계정으로 3,000핑이 들어올 예정입니다.")
                await message.channel.send( gid + "에서 출석체크 완료되었습니다.")
                file1 = openpyxl.load_workbook('\sheets\enchant.xlsx')
                fin = False
                sheet1 = file1.active
                for b in range(1, 51):
                    if sheet1["C" + str(b)].value == message.author.name:
                        for x in range(1, 3):
                            xy = random.choice(["A", "B", "C", "D"])
                            if xy == "A":
                                sheet["B" + str(b)].value = int(sheet["B" + str(b)].value) + 3
                                message.channel.send("운이 좋네요! 자신의 부서진 아이템을 제외한 맨 처음으로 생성한 아이템의 레벨이  ***3***  만큼 올랐습니다!")
                                fin = True
                                break
                    if fin == False:
                        sheet["B" + str(i)].value = int(sheet["B" + str(b)].value) + 1
                        message.channel.send("자신의 부서진 아이템을 제외한 맨 처음으로 생성한 아이템의 레벨이  ***1***  만큼 올랐습니다!")



                await message.channel.send()
                break
        file.save('\sheets\chullcheck.xlsx')
    if message.content.startswith("pg!출석정보만들기"):
        gid = message.author.guild.name
        file = openpyxl.load_workbook('\sheets\chullcheck.xlsx')
        sheet = file.active
        for i in range(1, 151):
            if sheet["A" + str(i)].value == message.author.name and sheet["E" + str(i)].value == gid:
                await message.channel.send("이미 정보가 만들어진 계정입니다.")
                break
            else:
                if sheet["A" + str(i)].value == "-":
                    sheet["A" + str(i)].value = message.author.name
                    sheet["E" + str(i)].value = str(gid)
                    await message.channel.send(gid + "에서 정상적으로 제작되었습니다.")
                    break
        file.save('\sheets\chullcheck.xlsx')
    if message.content.startswith("pg!출순"):
        learn = message.content.split(" ")
        file = openpyxl.load_workbook('\sheets\chullcheck.xlsx')
        sheet = file.active
        for i in range(1, 151):
            if sheet["A" + str(i)].value == learn[1] and sheet["E" + str(i)].value == str(gid):
                embed4 = discord.Embed(title=learn[1] + " 님 출첵 개요", color=0xff0000)
                embed4.add_field(name="닉네임", value=learn[1], inline=False)
                embed4.add_field(name="맨 마지막으로 출석 체크한 날짜(월일)", value=sheet["C" + str(i)].value, inline=False)
                embed4.add_field(name="출석체크한 수", value=sheet["B" + str(i)].value, inline=False)
                embed4.set_footer(text="먀우 많이 사랑해주세요!")
                await message.channel.send(embed=embed4)
    # 강화 관련 코드
    if message.content.startswith("pg!강템생성"):
        file = openpyxl.load_workbook('\sheets\enchant.xlsx')
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == learn[1]:
                await message.channel.send("이미 존재하는 아이템이예요!")
                break
            if sheet["A" + str(i)].value == "-":
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = "1"
                sheet["C" + str(i)].value = message.author.name
                await message.channel.send("성공적으로 생성했어요!")
                break
        file.save('\sheets\enchant.xlsx')

    if message.content.startswith("pg!강화"):
        file = openpyxl.load_workbook('\sheets\enchant.xlsx')
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 51):
            if message.author.name == sheet["C" + str(i)].value == message.author.name and sheet["A" + str(i)].value == learn[1]:
                enchant = 50 - sheet["B" + str(i)].value
                for a in range(1, enchant):
                    success = False
                    randomvar = random.choice(["A", "B", "C", "D", "E", "F"])
                    if sheet["B" + str(i)].value == "50":
                        await message.channel.send("더 이상 강화할 수 없습니다.")
                        success = "None"
                        break
                    if randomvar == "A":
                        success = True
                        break
                if success == True:
                    await message.channel.send(sheet["A" + str(i)].value + "(이)가 강화에 ***성공***했어요!.")
                    yrank = sheet["B" + str(i)].value
                    sheet["B" + str(i)].value = int(sheet["B" + str(i)].value) + 1
                    await message.channel.send(
                        sheet["A" + str(i)].value + "(이)가 " + yrank + "에서 " + sheet["B" + str(i)].value + "까지 올랐어요!")
                if success == False:
                    await message.channel.send("어이쿠, 이런! 아이템이 부서졌어요!")
                    file.active.delete_rows(int(i))
                    sheet["A" + str(i)].value = "-"
        file.save('\sheets\enchant.xlsx')
    if message.content.startswith("pg!견적서"):
        file = openpyxl.load_workbook('\sheets\enchant.xlsx')
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 51):
            if learn[1] == sheet["A" + str(i)].value and message.author.name != sheet["C" + str(i)].value:
                await message.channel.send("언급한 아이템이 본인 것이 아닙니다.")
            if learn[1] == sheet["A" + str(i)].value and message.author.name == sheet["C" + str(i)].value:
                embed3 = discord.Embed(title="견적서", color=0xff0000)
                embed3.add_field(name="이름", value=sheet["A" + str(i)].value, inline=False)
                embed3.add_field(name="현 등급", value="lv. " + str(sheet["B" + str(i)].value), inline=False)
                embed3.add_field(name="수리사", value=sheet["C" + str(i)].value, inline=False)
                embed3.set_footer(text="먀우 많이 사랑해주세요!")
                await message.channel.send(embed=embed3)
        # 강화 관련 끝
    if message.content.startswith("pg!역할지급 19"):
        id = message.author.id
        role = discord.utils.get(message.guild.roles, name="| 19 [ 19 ] |")
        await message.author.add_roles(role)
        await message.channel.send("<@" + str(id) + "> 님,성공적으로 지급되었습니다! 판도라의 상자가 열렸군요.....")

    if message.content.startswith("pg!역할제거 구독자"):
        id = message.author.id
        role = discord.utils.get(message.guild.roles, name="| 🎁구독자 |")
        await message.author.remove_roles(role)
        await message.channel.send("<@" + str(id) + "> 님,성공적으로 제거되었습니다")

    if message.content.startswith("pg!역할제거 19"):
        id = message.author.id
        role = discord.utils.get(message.guild.roles, name="| 19 [ 19 ] |")
        await message.author.remove_roles(role)
        await message.channel.send("<@" + str(id) + "> 님,성공적으로 제거되었습니다")
    if message.content.startswith("pg!가위바위종이"):
        iserror = False
        try:
            loadplayer = message.content[10:12]
        except discord.errors.HTTPException:
            await message.channel.send("허어;; 뭘 잘못 쓰신 거 같습니다만;;")
            iserror = True
        finished = False
        aigab = random.choice(["가위", "바위", "종이"])
        print("가위바위보 ai 값" + aigab)
        if iserror == False:
            print("가위바위보 플레이어 값" + loadplayer)
            if loadplayer == "가위":
                playergab = "가위"
                finished = True
            if loadplayer == "바위":
                playergab = "바위"
                finished = True
            if loadplayer == "종이":
                playergab = "종이"
                finished = True
            if finished == False:
                await message.channel.send("허어;; 뭘 잘못 쓰신 거 같습니다만;;")
                iserror = True
        if iserror == False:
            if aigab == "가위":
                # ai값이 가위일때(차례대로, 비기고, 지고, 이기고)
                if playergab == "가위":
                    await message.channel.send("플레이어는 가위를 냈고.....!")
                    time.sleep(1)
                    await message.channel.send("먀우는 가위를 냈습니다! ***비겼습니다!***")
                if playergab == "종이":
                    await message.channel.send("플레이어는 종이를 냈고.....!")
                    time.sleep(1)
                    await message.channel.send("먀우는 가위를 냈습니다! ***졌습니다!***")
                if playergab == "바위":
                    await message.channel.send("플레이어는 바위를 냈고.....!")
                    time.sleep(1)
                    await message.channel.send("먀우는 가위를 냈습니다! ***이겼습니다!***")
            if aigab == "바위":
                # ai값이 바위일때(차례대로, 비기고, 지고, 이기고)
                if playergab == "바위":
                    await message.channel.send("플레이어는 바위를 냈고.....!")
                    time.sleep(1)
                    await message.channel.send("먀우는 바위를 냈습니다! ***비겼습니다!***")
                if playergab == "가위":
                    await message.channel.send("플레이어는 가위를 냈고.....!")
                    time.sleep(1)
                    await message.channel.send("먀우는 바위를 냈습니다! ***졌습니다!***")
                if playergab == "종이":
                    await message.channel.send("플레이어는 종이를 냈고.....!")
                    time.sleep(1)
                    await message.channel.send("먀우는 바위를 냈습니다! ***이겼습니다!***")
            if aigab == "종이":
                # ai값이 종이일때(차례대로, 비기고, 지고, 이기고)
                if playergab == "종이":
                    await message.channel.send("플레이어는 종이를 냈고.....!")
                    time.sleep(1)
                    await message.channel.send("먀우는 종이를 냈습니다! ***비겼습니다!***")
                if playergab == "바위":
                    await message.channel.send("플레이어는 바위를 냈고.....!")
                    time.sleep(1)
                    await message.channel.send("먀우는 종이를 냈습니다! ***졌습니다!***")
                if playergab == "가위":
                    await message.channel.send("플레이어는 가위를 냈고.....!")
                    time.sleep(1)
                    await message.channel.send("먀우는 종이를 냈습니다! ***이겼습니다!***")
    # 가위바위보 끝
    # 학습 관련 코드
    if message.content.startswith("pg!배워"):
        file = openpyxl.load_workbook('\sheets\learn.xlsx')
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == learn[1]:
                await message.channel.send("이미 배웠어요!")
                break
            if sheet["A" + str(i)].value == "-":
                sheet["A" + str(i)].value = learn[1]
                sheet["B" + str(i)].value = learn[2]
                sheet["C" + str(i)].value = "***by***" + "  " + message.author.name
                sheet["D" + str(i)].value = message.author.name
                await message.channel.send("성공적으로 배웠어요!")
                break
    if message.content.startswith("pg!수정"):
        file = openpyxl.load_workbook('\sheets\learn.xlsx')
        sheet = file.active
        learn = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == learn[1] and sheet["D" + str(i)].value == message.author.name:
                sheet["B" + str(i)].value = learn[2]
                await message.channel.send("성공적으로 수정했어요!")
        file.save('\sheets\learn.xlsx')
    if message.content.startswith("pg!배운거말해"):
        file = openpyxl.load_workbook('\sheets\learn.xlsx')
        sheet = file.active
        memory = message.content.split(" ")
        for i in range(1, 51):
            if sheet["A" + str(i)].value == memory[1]:
                await message.channel.send(sheet["B" + str(i)].value + "  " + sheet["C" + str(i)].value)
                break
    # 학습 관련 코드 끝
    if message.content.startswith("pg!파도"):
        # 파도 기능 구현용 변수 설정
        messagelen = len(message.content)
        wavemessage = message.content[6:int(messagelen)]
        messagelen1 = int(len(wavemessage))
        i = 1
        x = 7
        # 미친구조
        roop = True
        while i<=messagelen1:
            # 대량 도배 방지 코드(우려먹기)
            if messagelen1>=20:
                await message.channel.send("메시지가 너무 깁니다")
                break
            await message.channel.send(message.content[6:x])
            x = x + 1
            i = i + 1
        # 세팅 한번만하고
        x = x - 2
        # 갑시다
        print(i)
        while i>=1:
            try:
                await message.channel.send(message.content[6:x])
                if messagelen1>=20:
                    raise discord.errors.HTTPException  # 아 귀찮아
            except discord.errors.HTTPException:
                await message.channel.send("출력 끝!")
                break
            x = x - 1
            i = i - 1
    if message.content.startswith("pg!도배"):
        messagelen = len(message.content)
        wavemessage = message.content[6:int(messagelen)]
        messagelen1 = int(len(wavemessage))
        i = 1
        while i<=messagelen:
            if messagelen1>=20:
                await message.channel.send("메시지가 너무 깁니다")
                break
            await message.channel.send(wavemessage)
            i = i + 1
        await message.channel.send("도배 완료!")

    #운빨겜 관련 코드
    if message.content.startswith("pg!운겜생성"):
        file = openpyxl.load_workbook('\sheets\luck.xlsx')
        sheet = file.active
        isbreak = False
        learn = message.content.split(" ")
        try:
            firstname = learn[1]
        except IndexError:
            message.channel.send("잘못 입력하셨습니다. `pg!운겜생성 [첫 번째 아바타]` 형식으로 입력하세요.")
            isbreak = True
        for i in range(1, 11):
            if isbreak == True:
                break
            if sheet[ "A" + str(i)].value == "-":
                sheet[ "A" + str(i)].value = firstname
                message.channel.send("운빨겜이 생성되었습니다. pg!운빨겜 " + firstname + " 로 " + firstname + "의 소유자님이 만든 운빨겜에 참가하세요!")
        file.save('\sheets\luck.xlsx')
    if message.content.startswith("pg!운빨겜"):
        file = openpyxl.load_workbook('\sheets\luck.xlsx')
        sheet = file.active
        isbreak = False
        learn = message.content.split(" ")
        try:
            firstname = learn[1]
        except IndexError:
            message.channel.send("잘못 입력하셨습니다. `pg!운빨겜 [내 아바타]` 형식으로 입력하세요.")
            isbreak = True
        for i in range(1, 11):
            if isbreak == True:
                break
            if sheet["B" + str(i)].value == "-":
                sheet["B" + str(i)].value = firstname
                message.channel.send(firstname + "이 두 번째 선수로 " + sheet["A" + str(i)].value + "님의 운빨겜에 참가했습니다!")
                isbreak = True
        for i in range(1, 11):
            if isbreak == True:
                break
            if sheet["C" + str(i)].value == "-":
                sheet["C" + str(i)].value = firstname
                message.channel.send(firstname + "이 세 번째 선수로 " + sheet["A" + str(i)].value + "님의 운빨겜에 참가했습니다!")
                isbreak = True
        for i in range(1, 11):
            if isbreak == True:
                break
            if sheet["D" + str(i)].value == "-":
                sheet["D" + str(i)].value = firstname
                message.channel.send(firstname + "이 네 (마지막) 번째 선수로 " + sheet["A" + str(i)].value + "님의 운빨겜에 참가했습니다!")
                message.channel.send("이제 운빨겜의 승자를 가립니다. 두구두구 ***두구두구?????***")
                winner = random.choice(["A", "B", "C", "D"])
                message.channel.send(sheet[str(winner) + str(i)].value + "님이 운빨겜에서 ***우승***하셨습니다!")
                file.active.delete_rows(int(i))
                sheet["A" + str(i)].value = "-"
    file.save('\sheets\luck.xlsx')




        



    # 여기서부턴 비밀코드 ㅇㅅㅇ
    if message.content.startswith("pg!도움마"):
        msg = await message.channel.send("***바보냐?***")
        time.sleep(1)
        await msg.edit(content="판사님 저는 아무 말도 하지 않았습니다")
    if message.content.startswith("ㅇㅅㅇ"):
        id = message.author.id
        await message.channel.send("<@" + str(id) + ">, 뀨?")
    if message.content.startswith("ㅇㅂㅇ"):
        id = message.author.id
        await message.channel.send("<@" + str(id) + ">, 흠.....")
    if message.content.startswith("ㄴㅇ0ㅇㄱ"):
        id = message.author.id
        await message.channel.send("<@" + str(id) + ">, 더 상상 못함 ㅇㅅㅇ")
    if message.content.startswith("ㄴ0ㄱ"):
        id = message.author.id
        await message.channel.send("<@" + str(id) + ">, 상상 못함 ㅇㅅㅇ")
    if message.content.startswith("ㄴㅇㄱ"):
        id = message.author.id
        await message.channel.send("<@" + str(id) + ">, 상상 못함 ㅇㅅㅇ")
    if message.content.startswith("ㅇㅁㅇ"):
        id = message.author.id
        await message.channel.send("<@" + str(id) + ">, [벙어리 모드 가동.]")


client.run("NjkwMzU4NjQzODgzMTE0NTQ2.XnQQ0A.TZZWRBtSYcOvN2uVNzDCtuDQdOg")
